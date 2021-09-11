from io import BytesIO

import openpyxl
from django.conf import settings
from django.core.files import File
from django.core.files.storage import default_storage
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views import generic
from openpyxl_image_loader import SheetImageLoader
from webpush import send_user_notification

from application.views import error_403
from .forms import *
from .models import *


# Create your views here.
# my superuser check decorator
def superuser(any_func):
    def inner(*args, **kwargs):
        if not args[0].user.is_superuser:
            return error_403(args[0], 'none')
        else:
            return any_func(args[0], **kwargs)

    return inner


@superuser
def index(request):
    webpush_settings = getattr(settings, 'WEBPUSH_SETTINGS', {})
    vapid_key = webpush_settings.get('VAPID_PUBLIC_KEY')
    user = request.user

    return render(request, 'adminPanel/admin_home.html', {user: user, 'vapid_key': vapid_key})


# all user
@superuser
def all_user(request):
    return render(request, 'adminPanel/all_user.html', {'list': User.objects.all()})


# kitchen list admin
@method_decorator(superuser, name='dispatch')
class ListKitchenView(generic.ListView):
    model = KitchenCategory
    template_name = 'adminPanel/admin_kitchen.html'
    context_object_name = 'kitchen_list'

    def get_context_data(self, *args, **kwargs):
        ctx = super(ListKitchenView, self).get_context_data(*args, **kwargs)
        ctx['product'] = 'Kitchen'
        return ctx


# kitchen detail
@method_decorator(superuser, name='dispatch')
class KitchenDetailview(generic.ListView):
    model = Kitchen
    template_name = 'adminPanel/admin_kitchen_list_detail.html'
    context_object_name = 'kitchen_detail'

    def get_queryset(self):
        kitchen_category = KitchenCategory.objects.filter(pk=self.kwargs['pk'])
        return Kitchen.objects.select_related('kitchen_type').filter(kitchen_type=kitchen_category[0])


# kitchen add
@superuser
def kitchenadd(request):
    ctx = {'form': KitchenForm}
    if request.method == 'POST':
        form = KitchenForm(request.POST or None, request.FILES or None)

        if form.is_valid():
            form.save(commit=True)
            return redirect('adminPanel:admin-kitchen')

    return render(request, 'adminPanel/admin_add_kitchen.html', ctx)


# complete kitchen detail
@method_decorator(superuser, name='dispatch')
class CompleteKitchenDetail(generic.DetailView):
    model = Kitchen
    template_name = 'adminPanel/admin_kitchen_detail.html'
    context_object_name = 'kitchen_detail'

    def get_context_data(self, **kwargs):
        ctx = super(CompleteKitchenDetail, self).get_context_data(**kwargs)
        return ctx


# units for kitchen
@method_decorator(superuser, name='dispatch')
class UnitsList(generic.ListView):
    model = Units
    paginate_by = 10
    template_name = 'adminPanel/units.html'
    context_object_name = 'units'

    def get_context_data(self, *args, **kwargs):
        ctx = super(UnitsList, self).get_context_data(*args, **kwargs)
        ctx['kitchen'] = Kitchen.objects.select_related('kitchen_type').get(pk=self.kwargs['pk'])
        return ctx

    def get_queryset(self):
        kitchen = Kitchen.objects.select_related('kitchen_type').get(pk=self.kwargs['pk'])
        return Units.objects.select_related('kitchen').filter(kitchen=kitchen.kitchen_type)


@superuser
def addunitcategory(request):
    if request.method == 'POST':
        UnitType.objects.create(name=request.POST['category'])
        return redirect('adminPanel:admin-add-units')


@superuser
def addUnits(request, **kwargs):
    ctx = {'form': AddUnitsForm}
    if request.method == 'POST':
        if not kwargs:
            form = AddUnitsForm(request.POST or None, request.FILES or None)
            if form.is_valid():
                data=form.cleaned_data
                form.save(commit=True)
                return redirect(
                    reverse_lazy('adminPanel:units', kwargs={'pk': KitchenCategory.objects.get(name=data['kitchen']).pk}))
        else:
            unit = Units.objects.get(pk=kwargs['pk'])
            img = unit.img
            unit.name = request.POST['name']
            unit.description = request.POST['description']
            unit.price = request.POST['price']
            unit.kitchen = KitchenCategory.objects.get(pk=request.POST['kitchen'])
            unit.unit_type = UnitType.objects.get(pk=request.POST['type'])
            if request.FILES.get('img') is not None:
                unit.img = request.FILES.get('img')
                unit.save()
            else:
                unit.img = img
                unit.save()

    if kwargs:
        ctx['item'] = Units.objects.select_related('unit_type').get(pk=kwargs['pk'])
        ctx['type'] = UnitType.objects.all()
        ctx['s_type'] = ctx['item'].unit_type
        ctx['kitchen'] = Kitchen.objects.select_related('kitchen_type').all()
        ctx['s_kitchen'] = ctx['item'].kitchen
        ctx['name'] = 'name'

    return render(request, 'adminPanel/add_units.html', ctx)


# worktop category
@method_decorator(superuser, name='dispatch')
class WorkTypeCategory(generic.ListView):
    model = Worktop_category
    template_name = 'adminPanel/admin_worktops.html'
    context_object_name = 'category'

    def get_context_data(self, *args, **kwargs):
        ctx = super(WorkTypeCategory, self).get_context_data(*args, **kwargs)
        ctx['product'] = 'Worktop'
        return ctx


# add worktop category
@superuser
def add_worktop_category(request, **kwargs):
    if request.method == 'POST':
        if kwargs['name'] == 'Worktop':
            Worktop_category.objects.create(worktop_type=request.POST['Worktop'])
            return redirect('adminPanel:admin-worktop')
        elif kwargs['name'] == 'Appliances':
            Category_Applianes.objects.create(name=request.POST['Appliances'])
            return redirect('adminPanel:admin-appliances')

        elif kwargs['name'] == 'Kitchen':
            KitchenCategory.objects.create(name=request.POST['Kitchen'])
            return redirect('adminPanel:admin-kitchen')


# worktop list on category basis
@method_decorator(superuser, name='dispatch')
class WorktopList(generic.ListView):
    model = WorkTop
    paginate_by = 10
    template_name = 'adminPanel/admin_worktop_list.html'
    context_object_name = 'list'

    def get_context_data(self, *args, **kwargs):
        ctx = super(WorktopList, self).get_context_data(*args, **kwargs)
        ctx['product'] = 'Worktop'
        return ctx

    def get_queryset(self):
        worktop_category = Worktop_category.objects.get(pk=self.kwargs['pk'])
        return WorkTop.objects.select_related('category').filter(category=worktop_category)


# worktop detail
@method_decorator(superuser, name='dispatch')
class WorktopDetail(generic.DetailView):
    model = WorkTop
    template_name = 'adminPanel/admin_worktops_detail.html'
    context_object_name = 'detail'

    def get_context_data(self, **kwargs):
        ctx = super(WorktopDetail, self).get_context_data(**kwargs)
        ctx['product'] = 'Worktop'
        return ctx

    def get_queryset(self):
        return WorkTop.objects.prefetch_related('category').filter(pk=self.kwargs['pk'])


# adding/editing worktop
@superuser
def add_worktop(request, **kwargs):
    # get request to display form
    if kwargs['name'] == 'Worktop':
        ctx = {'form': WorktopForm, 'product': 'Worktop'}
    else:
        ctx = {'form': AppliancesForm, 'product': 'Appliances'}

    if request.method == 'POST':

        # worktop
        if kwargs['name'] == 'Worktop':
            if len(kwargs) == 1:
                category = Worktop_category.objects.get(pk=request.POST['category'])
                form = WorktopForm(request.POST, request.FILES)
                if form.is_valid():
                    form.save(commit=True)
                    return redirect(reverse_lazy('adminPanel:admin-worktop-list', kwargs={'pk': category.pk}))

            else:
                category = Worktop_category.objects.get(pk=request.POST['category'])

                worktop_instance = WorkTop.objects.get(pk=kwargs['pk'])
                img = worktop_instance.worktop_img
                worktop_instance.category = category
                worktop_instance.name = request.POST['name']
                worktop_instance.color = request.POST['name']
                worktop_instance.description = request.POST['description']
                worktop_instance.price = request.POST['price']
                worktop_instance.size = request.POST['size']
                if request.FILES.get('img') is None:
                    worktop_instance.worktop_img = img
                else:
                    worktop_instance.worktop_img = request.FILES.get('img')
                worktop_instance.save()

                return redirect(reverse_lazy('adminPanel:admin-worktop-list', kwargs={'pk': category.pk}))


        # appliances
        else:

            if len(kwargs) == 1:
                form = AppliancesForm(request.POST, request.FILES)
                if form.is_valid():
                    form.save(commit=True)
                    category = Category_Applianes.objects.get(pk=request.POST['category'])
                    return redirect(reverse_lazy('adminPanel:admin-appliances-list', kwargs={'pk': category.pk}))

            else:
                category = Category_Applianes.objects.filter(pk=request.POST['category'])
                price = float(request.POST['price'])
                appliance_instance = Appliances.objects.get(pk=kwargs['pk'])
                img = appliance_instance.img
                appliance_instance.category.pk = category[0].pk
                appliance_instance.name = request.POST['name']
                appliance_instance.appliances_type = request.POST['type']
                appliance_instance.description = request.POST['description']
                appliance_instance.price = price
                appliance_instance.brand_name = request.POST['brand']
                if request.FILES.get('img') is None:
                    appliance_instance.img = img
                else:
                    appliance_instance.img = request.FILES.get('img')
                appliance_instance.save()
                return redirect(reverse_lazy('adminPanel:admin-appliances-list', kwargs={'pk': category[0].pk}))

    # get request with kwargs for or if editing a existing value
    if kwargs and len(kwargs) != 1:
        if kwargs['name'] == 'Appliances':
            ctx['add'] = 'add'
            ctx['item2'] = Appliances.objects.get(pk=kwargs['pk'])
            return render(request, 'adminPanel/admin_work_add.html', ctx)
        else:
            ctx['add2'] = 'add2'
            ctx['item'] = WorkTop.objects.get(pk=kwargs['pk'])
    return render(request, 'adminPanel/admin_work_add.html', ctx)


# appliances

# appliances category
@method_decorator(superuser, name='dispatch')
class AppliancesCategory(generic.ListView):
    model = Category_Applianes
    template_name = 'adminPanel/admin_worktops.html'
    context_object_name = 'category'

    def get_context_data(self, *args, **kwargs):
        ctx = super(AppliancesCategory, self).get_context_data(*args, **kwargs)
        ctx['product'] = 'Appliances'
        return ctx


# appliances list based on category
@method_decorator(superuser, name='dispatch')
class AppliancesList(generic.ListView):
    model = Appliances
    template_name = 'adminPanel/admin_worktop_list.html'
    context_object_name = 'list'

    def get_context_data(self, *args, **kwargs):
        ctx = super(AppliancesList, self).get_context_data(*args, **kwargs)
        ctx['product'] = 'Appliances'
        return ctx

    def get_queryset(self):
        category = Category_Applianes.objects.get(pk=self.kwargs['pk'])
        return Appliances.objects.prefetch_related('category').filter(category=category)


# appliances detail
@method_decorator(superuser, name='dispatch')
class AppliancesDetail(generic.DetailView):
    model = Appliances
    template_name = 'adminPanel/admin_worktops_detail.html'
    context_object_name = 'detail'

    def get_context_data(self, **kwargs):
        ctx = super(AppliancesDetail, self).get_context_data(**kwargs)
        ctx['product'] = 'Appliances'
        return ctx


# Carts

# All orders
@method_decorator(superuser, name='dispatch')
class Orders(generic.ListView):
    model = CompleteOrder
    template_name = 'adminPanel/admin_complete_orders.html'
    context_object_name = 'list'

    def get_queryset(self):
        return CompleteOrder.objects.prefetch_related('order', 'user')


# detail order view
@method_decorator(superuser, name='dispatch')
class DetailOrder(generic.DetailView):
    model = CompleteOrder
    template_name = 'adminPanel/admin_order_detail.html'
    context_object_name = 'detail'

    def get_context_data(self, **kwargs):
        ctx = super(DetailOrder, self).get_context_data(**kwargs)
        cart = CompleteOrder.objects.select_related('user').get(pk=self.kwargs['pk'])
        price = 0
        sample_price = 0
        for item in cart.order.all():
            if item.worktop:

                if item.sample_worktop == 'Yes':
                    sample_price += 5
                else:
                    price += item.worktop.price * item.qty
            if item.appliances:
                price += item.appliances.price * item.qty
            if item.accessories:
                price += item.accessories.price * item.qty
            if item.kitchen_order:
                if item.kitchen_order.units_intermediate_set.all():
                    for i in item.kitchen_order.units_intermediate_set.all():
                        price += i.unit.price * i.qty
                else:
                    sample_price += 4.99
            if item.service:
                sample_price += 0
        if price < 300 and price != 0:
            price += 30
        price = price + sample_price
        ctx['total_price'] = price
        return ctx


@method_decorator(superuser, name='dispatch')
class BlogsList(generic.ListView):
    model = Blogs
    template_name = 'adminPanel/admin_blogs.html'
    context_object_name = 'list'


@superuser
def create_blog(request, **kwargs):
    ctx = {'form': BlogForm}

    if len(kwargs) != 1:
        blog = get_object_or_404(Blogs, id=kwargs['pk'])
        form = BlogForm(request.POST or None, request.FILES or None, instance=blog)
        ctx['form'] = form
        ctx['blog'] = blog
    if request.method == 'POST':
        if kwargs['to'] == 'add':
            form = BlogForm(request.POST, request.FILES)
            if form.is_valid():
                form.save(commit=True)
                return redirect('adminPanel:admin-blog')

        else:
            blog = get_object_or_404(Blogs, id=kwargs['pk'])
            form = BlogForm(request.POST or None, request.FILES or None, instance=blog)
            if form.is_valid():
                if request.FILES:
                    form.save(commit=True)
                else:
                    form.title_img = blog.title_img
                    form.save(commit=True)
                return redirect('adminPanel:admin-blog')
    return render(request, 'adminPanel/blog_add.html', ctx)


@method_decorator(superuser, name='dispatch')
class DetailBlog(generic.DetailView):
    model = Blogs
    template_name = 'adminPanel/blog-detail.html'
    context_object_name = 'detail'


@superuser
def bulk_add(request):
    form = FileForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save(commit=True)
            return file_reading(request)

    return render(request, 'adminPanel/testupload.html', {'form': form})


@superuser
def file_reading(request):
    file = UploadFile.objects.first()
    excel_sheet = openpyxl.load_workbook(default_storage.open(file.file.name))
    if request.POST.get('name') == 'appliance':
        worksheet = excel_sheet[request.POST['category']]
        cat = Category_Applianes.objects.get(name=request.POST['category'])
        for i in range(3, worksheet.max_row):
            if worksheet.cell(row=i, column=3).value is not None:
                img_loader = SheetImageLoader(worksheet)
                img = img_loader.get(f'G{i}')
                img_io = BytesIO()
                new_img = File(img_io, name=f'{cat}_{i}')
                if img.mode in 'RGBA' or img.mode in 'P':
                    img = img.convert('RGB')
                img.save(img_io, 'JPEG', optimize=True)
                app = Appliances.objects.create(
                    name=worksheet.cell(row=i, column=3).value,
                    appliance_category=worksheet.cell(row=i, column=5).value,
                    category=cat,
                    brand_name=worksheet.cell(row=i, column=4).value,
                    description=worksheet.cell(row=i, column=6).value,
                    img=new_img,
                    appliances_type=worksheet.cell(row=i, column=8).value,
                    price=worksheet.cell(row=i, column=9).value,
                    meta_name=worksheet.cell(row=i, column=10).value,
                    meta_title=worksheet.cell(row=i, column=11).value,
                    meta_description=worksheet.cell(row=i, column=12).value,

                )
                app.save()
        file.delete()
        return redirect('adminPanel:index')

    elif request.POST.get('name') == 'worktop':
        sheetname = excel_sheet.sheetnames
        worksheet = excel_sheet[sheetname[0]]
        cat = Worktop_category.objects.get(worktop_type__iexact=request.POST['category'])
        for i in range(3, worksheet.max_row):
            if worksheet.cell(row=i, column=3).value is not None:

                img_loader = SheetImageLoader(worksheet)
                img = img_loader.get(f'I{i}')
                img_io = BytesIO()
                new_img = File(img_io, name=f'{cat}_{i}')
                if img.mode in 'RGBA':
                    img = img.convert('RGB')
                img.save(img_io, 'JPEG', optimize=True)
                app = WorkTop.objects.create(
                    name=worksheet.cell(row=i, column=4).value,
                    category=cat,
                    color=worksheet.cell(row=i, column=5).value,
                    size=worksheet.cell(row=i, column=6).value,
                    description=worksheet.cell(row=i, column=7).value,
                    worktop_img=new_img,
                    price=worksheet.cell(row=i, column=8).value,
                    meta_name=worksheet.cell(row=i, column=10).value,
                    meta_title=worksheet.cell(row=i, column=11).value,
                    meta_description=worksheet.cell(row=i, column=12).value,
                )
                app.save()
        file.delete()
        return redirect('adminPanel:index')

    elif request.POST.get('name') == 'kitchen':
        sheetname = excel_sheet.sheetnames
        worksheet = excel_sheet[sheetname[0]]
        cat = KitchenCategory.objects.get(name__iexact=request.POST['category'])
        for i in range(8, worksheet.max_row):
            if worksheet.cell(row=i, column=3).value is not None:
                unit_cat = UnitType.objects.get_or_create(name=worksheet.cell(row=i, column=4).value)
                img_loader = SheetImageLoader(worksheet)
                img = img_loader.get(f'G{i}')
                img_io = BytesIO()
                if img.mode in 'P':
                    img = img.convert('RGBA')
                img.save(img_io, 'PNG', optimize=True)
                new_img = File(img_io, name=f'{cat}_{unit_cat}_{i}')
                app = Units.objects.create(
                    name=worksheet.cell(row=i, column=5).value,
                    unit_type=unit_cat[0],
                    kitchen=cat,
                    description=worksheet.cell(row=i, column=6).value,
                    price=worksheet.cell(row=i, column=8).value,
                    sku=worksheet.cell(row=i, column=9).value,
                    img=new_img
                )
                app.save()
        file.delete()
        return redirect('adminPanel:index')

    elif request.POST.get('name') == 'accessory':
        sheetname = excel_sheet.sheetnames
        worksheet = excel_sheet[sheetname[0]]
        accessory_type = AccessoriesType.objects.get(name__iexact=request.POST['category'])
        for i in range(3, worksheet.max_row):
            if worksheet.cell(row=i, column=3).value is not None:

                img_loader = SheetImageLoader(worksheet)
                img = img_loader.get(f'E{i}')
                img_io = BytesIO()
                new_img = File(img_io, name=f'{accessory_type}_{i}')
                if img.mode in 'RGBA' or img.mode in 'P':
                    img = img.convert('RGB')
                img.save(img_io, 'JPEG', optimize=True)
                accessory = Accessories.objects.create(
                    accessories_type=accessory_type,
                    description=worksheet.cell(row=i, column=4).value,
                    img=new_img,
                    price=worksheet.cell(row=i, column=6).value,
                    sku=worksheet.cell(row=i, column=3).value,
                    meta_name=worksheet.cell(row=i, column=7).value,
                    meta_title=worksheet.cell(row=i, column=8).value,
                    meta_description=worksheet.cell(row=i, column=9).value,
                )
                accessory.save()
        file.delete()
        return redirect('adminPanel:index')
    return render(request, 'adminPanel/file.html')


@superuser
def approve(request):
    ctx = {'comments': Review.objects.select_related('appliances', 'worktop').filter(approval='Pending')}

    return render(request, 'adminPanel/comments.html', ctx)


def approving_admin(request, **kwargs):
    if kwargs['name'] == 'worktop':
        review = Review.objects.select_related('worktop', 'appliances').get(pk=kwargs['pk'])
        review.approval = 'Approved'
        review.save()
    else:
        review = Review.objects.select_related('worktop', 'appliances').get(pk=kwargs['pk'])
        review.approval = 'Approved'
        review.save()

    return redirect('adminPanel:approve')


def zip_qs_date(qs):
    date_list = []
    for item in qs:
        split_content = item.detail.split('/')
        if split_content:
            date_list.append(split_content[-1])
    return zip(qs, date_list)


@method_decorator(superuser, name='dispatch')
class ContactUsList(generic.ListView):
    model = ContactUs
    template_name = 'adminPanel/admin_contact_us_table.html'
    context_object_name = 'list'

    def get_context_data(self, **kwargs):
        ctx = super(ContactUsList, self).get_context_data(**kwargs)

        ctx['date'] = zip_qs_date(self.get_queryset())
        return ctx


@method_decorator(superuser, name='dispatch')
class ContactUsDetail(generic.DetailView):
    model = ContactUs
    template_name = 'adminPanel/admin_panel_contact_detail.html'
    context_object_name = 'detail'


@method_decorator(superuser, name='dispatch')
class ContactActualList(generic.ListView):
    model = ContactActual
    template_name = 'adminPanel/admin_actual_contact.html'
    context_object_name = 'list'

    def get_context_data(self, **kwargs):
        ctx = super(ContactActualList, self).get_context_data(**kwargs)

        ctx['date'] = zip_qs_date(self.get_queryset())
        return ctx


@method_decorator(superuser, name='dispatch')
class ContactActualDetail(generic.DetailView):
    model = ContactActual
    template_name = 'adminPanel/admin_actual_detail.html'
    context_object_name = 'detail'

    def get_context_data(self, **kwargs):
        return super(ContactActualDetail, self).get_context_data(**kwargs)


@method_decorator(superuser, name='dispatch')
class AccessoriesList(generic.ListView):
    model = AccessoriesType
    template_name = 'adminPanel/admin_accessories.html'
    context_object_name = 'list'

    def get_context_data(self, *args, **kwargs):
        return super(AccessoriesList, self).get_context_data(*args, **kwargs)


@method_decorator(superuser, name='dispatch')
class AccessoriesDetail(generic.ListView):
    template_name = 'adminPanel/admin_accessory_list.html'
    context_object_name = 'list'

    def get_queryset(self):
        return Accessories.objects.prefetch_related('accessories_type').filter(accessories_type_id=self.kwargs['pk'])


@method_decorator(superuser, name='dispatch')
class AccessoriesInDetail(generic.DetailView):
    model = Accessories
    template_name = 'adminPanel/admin_accessories_detail.html'
    context_object_name = 'detail'

    def get_context_data(self, **kwargs):
        return super(AccessoriesInDetail, self).get_context_data(**kwargs)


@superuser
def add_accessories(request, **kwargs):
    form = AccessoriesForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            form_instance = form.cleaned_data
            accessory = Accessories.objects.create(
                accessories_type=form_instance['accessories_type'],
                description=form_instance['description'],
                img=form_instance['img'],
                price=form_instance['price'],
                sku=form_instance['sku']
            )
            # form.save(commit=True)
            return redirect(
                reverse_lazy('adminPanel:detail_accessories', kwargs={'pk': form_instance.accessories_type.pk}))
    ctx = {'form': form}
    return render(request, 'adminPanel/admin_add_aaccessories.html', ctx)


@superuser
def edit_accessories(request, **kwargs):
    accessory_instance = Accessories.objects.select_related('accessories_type').get(pk=kwargs['pk'])
    form = AccessoriesForm(instance=accessory_instance)
    if request.method == 'POST':
        accessory_type = AccessoriesType.objects.get(pk=request.POST['accessories_type'])
        accessory_instance.accessories_type = accessory_type
        accessory_instance.description = request.POST['description']
        accessory_instance.price = request.POST['price']
        accessory_instance.sku = request.POST['sku']
        img = accessory_instance.img
        if request.FILES:
            img = request.FILES['img']
        accessory_instance.img = img
        accessory_instance.save()
        return redirect('adminPanel:detail_in_accessories', accessory_instance.accessories_type.pk,
                        accessory_instance.pk)

    ctx = {'form': form, 'accessory_instance': accessory_instance}
    return render(request, 'adminPanel/admin_edit_accessories.html', ctx)


@superuser
def add_accessories_type(request):
    if request.method == 'POST':
        AccessoriesType.objects.create(name=request.POST['name'])

        return redirect('adminPanel:accessories')


@method_decorator(superuser, name='dispatch')
class BrochureRequests(generic.ListView):
    model = Brochure
    template_name = 'adminPanel/admin_brochure.html'
    context_object_name = 'list'


def send_push(request):
    if request.method == 'POST':
        try:
            data = request.POST['text']

            user = request.user
            payload = {'head': request.POST['title'], 'body': data}
            send_user_notification(user=user, payload=payload, ttl=1000)

            return JsonResponse(status=200, data={"message": "Web push notifications successful"})
        except TypeError:
            return JsonResponse(status=500, data={"message": "An error occurred"})
